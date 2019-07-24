from numpy.f2py.cfuncs import includes
from openpyxl import Workbook
from openpyxl import load_workbook
from builtins import str
import numpy as np
from openpyxl.utils.exceptions import SheetTitleException
import xlrd
import pandas as  pd

f=open('RouteLibrary.txt','w')


wb= load_workbook('D:\SCU_BMW\MakeFiles\Generator\Generator.xlsm')
   

sheet=wb['RouteLibrary']






f.write("// ################################## RouteLibrary ################################## ")
f.write("\n")
def RouteLibrary(sheet):
   q=[]
   q2=[]
   q3=[]
   # insert all elements from column 2 in a list
   for row_cells in sheet.iter_rows(min_col=2, max_col=2):
      q.append(row_cells[0].value)
   
   
   for row_cells in sheet.iter_rows(min_col=1,max_col=1):
      if (row_cells[0].value != None):
         q2.append(row_cells[0].value)
         func = row_cells[0].value
      else:
         q2.append(func)

    
   for row_cells in sheet.iter_rows(min_col=7,max_col=7):
      q3.append(row_cells[0].value)
       
     
   for i in range(len(q)):
      if q[i] != None and q[i] != "FunctionName" and q[i] != "XXX END OF DEFINE XXX":
         if (q3[i] == None):
            f.write("""define(`{0}',`CALL  "/TestSequences\Library\Functions\FNCT_{1}\{0}"($*)')""".format(q[i],q2[i]))
         else:
            f.write("""define(`{0}',`CALL {2} = "/TestSequences\Library\Functions\FNCT_{1}\{0}"($*)')""".format(q[i],q2[i],q3[i]))
         f.write("\n")
               



RouteLibrary(sheet)
f.close()





   
   
